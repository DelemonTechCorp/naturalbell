from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db.models import Min, Max, Count
import datetime
from django.contrib import messages
import pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.core.files.storage import default_storage
from django.conf import settings
from django.utils.crypto import get_random_string
from .models import *
from django.http import JsonResponse
from django.http import HttpResponse
from datetime import datetime
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
import random
from datetime import date, timedelta
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse,JsonResponse
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect
import openpyxl
from openpyxl.utils import get_column_letter
# Create your views here.


def base(request):
    return render(request,'base.html')

def base1(request):
    return render(request,'base1.html')
def export_patients_to_excel(request):
    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Patients Data'

    # Write the header row
    headers = ['SI.No', 'Patient ID', 'Patient Name', 'Age', 'Nationality', 'Phone Number', 'Status', 'FromDate']
    sheet.append(headers)

    # Adjust column widths
    column_widths = [5, 15, 25, 10, 25, 15, 15, 15]
    for i, column_width in enumerate(column_widths, 1):  # 1-based index for columns
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = column_width

    # Retrieve the patient data
    patients = Patient.objects.all()

    for index, patient in enumerate(patients, start=1):
        # Get the related booking model for reg_date (FromDate)
        booking_obj = booking.objects.filter(patientid=patient).first()
        from_date = booking_obj.reg_date if booking_obj else ''

        # Write patient data to the Excel sheet
        sheet.append([
            index,
            patient.patient_id,
            patient.name,
            patient.age,
            patient.email,
            patient.phone,
            patient.status.Status if patient.status else '',
            from_date
        ])

    # Format the 'FromDate' column as a date
    date_column = get_column_letter(8)  # 'FromDate' is in the 8th column
    for cell in sheet[date_column]:
        if cell.row != 1:  # Skip the header
            cell.number_format = 'YYYY-MM-DD'  # Or any other date format you prefer

    # Set the content type and the filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=patients_data.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response
def importform(request):
    return render(request,'main/importpatient.html')
def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        if excel_file and excel_file.name.endswith('.xlsx'):
            try:
                # Load Excel file into pandas DataFrame
                df = pd.read_excel(excel_file, dtype={})
            except Exception as e:
                return render(request, 'main/importpatient.html', {'error': f'Error reading Excel file: {e}'})

            try:
                # Begin transaction to ensure data consistency
                with transaction.atomic():
                    admin = Admin.objects.first()
                    doctor = Doctor.objects.first()  # Example: Assign the first available doctor
                    num = 0
                    for index, row in df.iterrows():
                        num = num+1
                        try:
                            print(num)
                            # Extract patient data from each row
                            patient_id = row['patient_id'] if pd.notna(row['patient_id']) else None
                            print(patient_id,'patient_id')
                            patient_name = row['name']
                            patient_gender = row['gender']
                            patient_age = row['age'] if pd.notna(row['age']) else 0
                            patient_phone = row['phone']

                            # Convert registration_date to the correct format
                            registration_date = row['reg_date']
                            if pd.notna(registration_date):
                                try:
                                    # Attempt to parse the date
                                    registration_date = pd.to_datetime(registration_date, errors='coerce').strftime('%Y-%m-%d')
                                except Exception as date_error:
                                    messages.error(request, f"Error processing date in row {index + 1}: {date_error}")
                                    continue  # Skip this row if there's a date issue
                            else:
                                registration_date = None

                            # registration_time = row['reg_time'] if pd.notna(row['reg_time']) else None
                            registration_time = row['reg_time'] if pd.notna(row['reg_time']) and row['reg_time'].strip() else None
                            if registration_time is not None:
                               try:
        # Attempt to parse the time
                                  registration_time = pd.to_datetime(registration_time, format='%H:%M', errors='coerce').time()
                               except Exception as time_error:
                                        messages.error(request, f"Error processing time in row {index + 1}: {time_error}")
                                        continue
                            status = row['status'] if pd.notna(row['status']) else 'Registered'
                            nationality = row['nationality'] if pd.notna(row['nationality']) else None

                            # Generate a unique patient_id if not provided
                            if not patient_id:
                                print(num)
                                patient_id = f"PAT{get_random_string(length=7).upper()}"
                                while Patient.objects.filter(patient_id=patient_id).exists():
                                    patient_id = f"PAT{get_random_string(length=7).upper()}"

                            # Get or create status object
                            status_obj, created = Status.objects.get_or_create(Status=status)

                            # Create or update the Patient instance
                            patient, created = Patient.objects.update_or_create(
                                patient_id=patient_id,  # Using phone as a unique identifier
                                defaults={
                                    'phone': patient_phone,
                                    'name': patient_name,
                                    'gender': patient_gender,
                                    'age': patient_age,
                                    'email': nationality,
                                    'admin': admin,
                                    'status': status_obj
                                }
                            )

                            # Avoid duplicate bookings
                            if not booking.objects.filter(
                                    patientid=patient,
                                    Doctor=doctor,
                                    reg_date=registration_date,
                                    reg_time=registration_time).exists():
                                # Create a Booking instance associated with this patient
                                booking.objects.create(
                                    patientid=patient,
                                    Doctor=doctor,
                                    reg_date=registration_date,
                                    reg_time=registration_time,
                                    admin=admin,
                                    status=status_obj
                                )

                        except Exception as e:
                            messages.error(request, f"Error processing row {index + 1}: {str(e)}")
                            continue  # Continue processing other rows

                messages.success(request, "Data imported successfully!")
                return redirect('patient')

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return render(request, 'main/importpatient.html')

    return render(request, 'main/importpatient.html')
# def upload_excel(request):
#     if request.method == 'POST' and request.FILES['excel_file']:
#         excel_file = request.FILES['excel_file']

#         if excel_file and excel_file.name.endswith('.xlsx'):
#             try:
#                 # Load Excel file into pandas DataFrame
#                 df = pd.read_excel(excel_file, dtype={})
#             except Exception as e:
#                 return render(request, 'main/importpatient.html', {'error': f'Error reading Excel file: {e}'})

#             try:
#                 # Begin transaction to ensure data consistency
#                 with transaction.atomic():
#                     admin = Admin.objects.first()
#                     doctor = Doctor.objects.first()  # Example: Assign the first available doctor
#                     num = 0
#                     for index, row in df.iterrows():
#                         num += 1
#                         try:
#                             print(num)
#                             # Extract patient data from each row
#                             patient_id = row['patient_id'] if pd.notna(row['patient_id']) else None
#                             print(patient_id, 'patient_id')
#                             patient_name = row['name']
#                             patient_gender = row['gender']
#                             patient_age = row['age'] if pd.notna(row['age']) else 0
#                             patient_phone = row['phone']

#                             # Convert registration_date to the correct format
#                             registration_date = row['reg_date']
#                             if pd.notna(registration_date):
#                                 try:
#                                     registration_date = pd.to_datetime(registration_date, errors='coerce').strftime('%Y-%m-%d')
#                                 except Exception as date_error:
#                                     messages.error(request, f"Error processing date in row {index + 1}: {date_error}")
#                                     continue  # Skip this row if there's a date issue
#                             else:
#                                 registration_date = None

#                             # Convert and validate registration_time
#                             registration_time = row['reg_time'] if pd.notna(row['reg_time']) else None
#                             if registration_time:
#                                 try:
#                                     # Parse the time to ensure it's in the correct format
#                                     registration_time = datetime.strptime(str(registration_time), '%H:%M').time()
#                                 except ValueError as time_error:
#                                     messages.error(request, f"Error processing time in row {index + 1}: {time_error}")
#                                     continue  # Skip this row if there's a time format issue
#                             else:
#                                 registration_time = None  # Set to None if reg_time is null or missing

#                             status = row['status'] if pd.notna(row['status']) else 'Booked'
#                             nationality = row['nationality'] if pd.notna(row['nationality']) else None

#                             # Generate a unique patient_id if not provided
#                             if not patient_id:
#                                 print(num)
#                                 patient_id = f"PAT{get_random_string(length=7).upper()}"
#                                 while Patient.objects.filter(patient_id=patient_id).exists():
#                                     patient_id = f"PAT{get_random_string(length=7).upper()}"

#                             # Get or create status object
#                             status_obj, created = Status.objects.get_or_create(Status=status)

#                             # Create or update the Patient instance
#                             patient, created = Patient.objects.update_or_create(
#                                 patient_id=patient_id,  # Using phone as a unique identifier
#                                 defaults={
#                                     'phone': patient_phone,
#                                     'name': patient_name,
#                                     'gender': patient_gender,
#                                     'age': patient_age,
#                                     'email': nationality,
#                                     'admin': admin,
#                                     'status': status_obj
#                                 }
#                             )

#                             # Avoid duplicate bookings
#                             if not booking.objects.filter(
#                                     patientid=patient,
#                                     Doctor=doctor,
#                                     reg_date=registration_date,
#                                     reg_time=registration_time).exists():
#                                 # Create a Booking instance associated with this patient
#                                 booking.objects.create(
#                                     patientid=patient,
#                                     Doctor=doctor,
#                                     reg_date=registration_date,
#                                     reg_time=registration_time,  # Can be None if time is missing
#                                     admin=admin,
#                                     status=status_obj
#                                 )

#                         except Exception as e:
#                             messages.error(request, f"Error processing row {index + 1}: {str(e)}")
#                             continue  # Continue processing other rows

#                 messages.success(request, "Data imported successfully!")
#                 return redirect('patient')

#             except Exception as e:
#                 messages.error(request, f"An error occurred: {str(e)}")
#                 return render(request, 'main/importpatient.html')

#     return render(request, 'main/importpatient.html')
#main folder
# def dashboard(request):
#     login_id = request.session.get('alid')
#     patient=Patient.objects.filter(admin__Lid_id=login_id)
#     patient=patient.count()
#     today = timezone.now().date()
#     today_created_count = booking.objects.filter(created_at__date=today).count()
#     context={
#         'patient':patient,
#         'today_created_count': today_created_count,
#     }
#     return render(request,'main/dashboard.html',context)

def dashboard(request):
    total_patients = Patient.objects.count()
    total_therapists = Therapist.objects.count()
    today = timezone.now().date()
    per_page = int(request.GET.get('per_page', 10))
    today_created_count = booking.objects.filter(created_at__date=today).count()
    patients = Schedule.objects.all()


    # patient_statuses = []
    # for patient in patients:
    #     schedules = Schedule.objects.filter(patientid=patient).order_by('-Date')
    #     if schedules.exists():
    #         latest_schedule = schedules.first()
    #         start_date = latest_schedule.Date
    #         therapist = latest_schedule.Therapist
    #         therapy = latest_schedule.Therapy
    #     else:
    #         start_date = None
    #         therapist = None
    #         therapy = None

    #     patient_statuses.append({
    #         'name': patient.name,
    #         'start_date': start_date,
    #         'status': patient.status,  # Assuming patient.status is a string field
    #         'therapist': therapist,
    #         'therapy': therapy
    #     })
    paginator = Paginator(patients, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'total_patients': total_patients,
        'total_therapists': total_therapists,
        'today_created_count': today_created_count,
        # 'patient_statuses': patient_statuses,
        'patients':patients
    }

    return render(request, 'main/dashboard.html', context)
def appointment(request):
    doctors = Doctor.objects.all()
    treatment=Treatment.objects.all()
    patient=Patient.objects.all()
    context = {
        'doctor': doctors,
        'treatment':treatment,
        'patient':patient
    }
    return render(request,'main/appointment.html',context)
def addDoctorForm(request):
    department=Department.objects.all()
    context={
        'department':department
    }
    return render(request,'main/addDoctor.html',context)
def addDoctor(request):
     login_id = request.session.get('alid')
     admin = Admin.objects.filter(Lid_id=login_id).first()
     firstname=request.POST.get('firstname', None)
     lastname=request.POST.get('lastname', None)
     department=request.POST.get('department', None)
     speciality=request.POST.get('speciality', None)
     phno=request.POST.get('phno', None)
     email=request.POST.get('email', None)
     location=request.POST.get('location', None)
     about=request.POST.get('about', None)
     gender=request.POST.get('gender', None)
     image=request.FILES.get('image', None)
     obdept = Department.objects.get(id=department)
     ob = Login()
     ob.username = email
     ob.password = firstname+'@c1d3'
     ob.type = 'doctor'
     ob.save()
     ob1=Doctor()
     ob1.admin=admin
     ob1.Lid=ob
     ob1.name=firstname+" "+lastname
     ob1.contact_no=phno
     ob1.email_id=email
     ob1.place=location
     ob1.gender=gender
     ob1.Speciality=speciality
     ob1.About=about
     ob1.Image=image
     ob1.Department=obdept
     subject = 'Your Nature Bells Login Credentials'
     message = f"Here are your login credentials to access your account:\n\n- **Username**: {ob.username}\n\n- **Password**: {ob.password}\n\nPlease ensure to keep this information secure and do not share it with anyone.\n\nIf you have any questions or need assistance with the onboarding process, our customer care team is here to support you. Feel free to reach out to us at customercare@gmail.com.\n\nThank you for being a valuable part of NatureBells. We look forward to achieving great success together!\n\nBest regards,\n\nThe NatureBells Team"
     from_email = 'quickfixsorter123@gmail.com'
     recipient_list = [email]

     send_mail(subject, message, from_email, recipient_list)
     ob1.save()
     return HttpResponse("<script>alert('Inserted successfully');window.location='/viewDoctor'</script>")
# def addAppointment(request):
#      login_id = request.session.get('alid')
#      admin = Admin.objects.filter(Lid_id=login_id).first()
#      firstname=request.POST.get('fname', None)
#      lastname=request.POST.get('lname', None)
#      dob=request.POST.get('dob', None)
#      appointmentdate=request.POST.get('apdate', None)
#      phno=request.POST.get('phno', None)
#      email=request.POST.get('email', None)
#      appointmenttime=request.POST.get('aptime', None)
#      about=request.POST['note']
#      gender=request.POST['gender']
#      doctors=request.POST.get('doctor', None)
#      treatment=request.POST.get('treatment',None)
#      obdoc = Doctor.objects.get(id=doctors)
#      obtreatment = Treatment.objects.get(id=treatment)
#      ob1=Patient()
#      ob1.admin=admin
#      ob1.name=firstname+" "+lastname
#      ob1.phone=phno
#      ob1.email=email
#      ob1.gender=gender
#      ob1.DOB=dob
#      ob1.note=about

#      ob1.doctor=obdoc
#      ob1.treatment=obtreatment
#      if dob:
#         dob_date = datetime.strptime(dob, "%Y-%m-%d").date()
#         today = datetime.today().date()
#         age = today.year - dob_date.year - ((today.month, today.day) < (dob_date.month, dob_date.day))
#      else:
#         age = None
#      ob1.age=age
#      booking_status, created = Status.objects.get_or_create(Status='Booked')
#      ob1.status = booking_status
#      ob1.save()
#      ob2=booking()
#      ob2.admin=admin
#      ob2.Doctor=obdoc
#      ob2.patientid=ob1
#      ob2.reg_date=appointmentdate
#      ob2.reg_time=appointmenttime
#      ob2.about=about
#      ob2.save()
#      return HttpResponse("<script>alert('Inserted successfully');window.location='/report'</script>")
def addAppointment(request):
    login_id = request.session.get('alid')
    admin = Admin.objects.filter(Lid_id=login_id).first()

    # Check if the selected patient is new or existing
    patient_id = request.POST.get('pid', None)
    if patient_id == "new":
        # New patient, create a new Patient object
        firstname = request.POST.get('fname', None)
        lastname = request.POST.get('lname', None)
        dob = request.POST.get('dob', None)
        phno = request.POST.get('phno', None)
        email = request.POST.get('email', None)
        about = request.POST.get('note', None)
        gender = request.POST.get('gender', None)
        treatment = request.POST.get('treatment', None)
        obtreatment = Treatment.objects.get(id=treatment)
        doctor = request.POST.get('doctor', None)
        obdoctor = Doctor.objects.get(id=doctor)
        ob1 = Patient(
            admin=admin,
            name=firstname + " " + lastname,
            phone=phno,
            email=email,
            gender=gender,
            DOB=dob,
            note=about,
            treatment=obtreatment,
            doctor=obdoctor
        )
        if dob:
            dob_date = datetime.strptime(dob, "%Y-%m-%d").date()
            today = datetime.today().date()
            age = today.year - dob_date.year - ((today.month, today.day) < (dob_date.month, dob_date.day))
        else:
            age = None
        ob1.age = age
        ob1.save()
    else:
        # Existing patient, retrieve the Patient object
        ob1 = Patient.objects.get(id=patient_id)

    # Continue with booking details
    appointmentdate = request.POST.get('apdate', None)
    appointmenttime = request.POST.get('aptime', None)
    doctors = request.POST.get('doctor', None)
    treatment = request.POST.get('treatment', None)
    obdoc = Doctor.objects.get(id=doctors)
    obtreatment = Treatment.objects.get(id=treatment)

    # Assign the retrieved objects to the booking
    ob2 = booking(
        admin=admin,
        Doctor=obdoc,
        patientid=ob1,
        reg_date=appointmentdate,
        reg_time=appointmenttime,
        about=request.POST.get('note', None),
        treatment=obtreatment
    )
    ob2.save()

    # Set the status to "Booked" for new bookings
    booking_status, created = Status.objects.get_or_create(Status='Registered')
    ob2.status = booking_status
    ob1.status=booking_status
    ob1.save()

    return HttpResponse("<script>alert('Inserted successfully');window.location='/report'</script>")


def Therapists(request):
    Therapies=Therapy.objects.all()
    treatment=Treatment.objects.all()
    context={
        'Therapies':Therapies,
        'treatment':treatment
    }
    return render(request,'main/Therapist.html',context)
def addTherapist(request):
     login_id = request.session.get('alid')
     admin = Admin.objects.filter(Lid_id=login_id).first()
     firstname=request.POST.get('fname', None)
     lastname=request.POST.get('lname', None)
    #  dob=request.POST.get('dob', None)
     speciality=request.POST.get('speciality', None)
     treatment=request.POST.get('treatment', None)
     phno=request.POST.get('phno', None)
     email=request.POST.get('email', None)
     location=request.POST.get('place', None)
     about=request.POST.get('about', None)
     gender=request.POST.get('gender', None)
    #  roomnumber=request.POST.get('roomnumber', None)
     image=request.FILES.get('image')
     obspe = Therapy.objects.get(id=speciality)
     obtreatment = Treatment.objects.get(id=treatment)
     ob1=Therapist()
     ob1.admin=admin
     ob1.name=firstname+" "+lastname
     ob1.contact_no=phno
     ob1.email_id=email
     ob1.place=location
     ob1.gender=gender
     ob1.specialization=obspe
     ob1.About=about
     ob1.Image=image
     ob1.treatment=obtreatment
    #  ob1.DOB=dob
    #  ob1.Room=roomnumber
     ob1.save()
     return HttpResponse("<script>alert('Inserted successfully');window.location='/viewTherapist'</script>")
def patient(request):
    search_query = request.GET.get('searchitem', '')
    per_page = int(request.GET.get('per_page', 10))
    login_id = request.session.get('alid')# Default to 10 if 'per_page' is not provided
    patient_list = Patient.objects.filter(admin__Lid_id=login_id).order_by('-id')
    patient_id = request.GET.get('patient_id')
    if patient_id:
        patient_list = patient_list.filter(id=patient_id)
    elif search_query:
        patient_list = patient_list.filter(name__icontains=search_query)


    paginator = Paginator(patient_list, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request,'main/patient.html',context)
def home(request):
    return render(request,'main/home.html')
def signin(request):
    return render(request,'main/signin.html')
def forgotpwd(request):
    return render(request,'main/forgot.html')

def getpwd(request):
    if request.method == 'POST':
        email_id = request.POST.get('email')  # Safely get the 'email' key

        if email_id:  # Check if email_id is not None or empty
            try:
                pwd = Login.objects.get(username=email_id)
                print(email_id, pwd)

                # Send the email
                send_mail(
                    'CREDITCARD FRAUD DETECTION',
                    "YOUR NEW PASSWORD IS  -" + pwd.password,
                    'email@gmail.com',
                    [email_id],
                    fail_silently=False
                )
                return HttpResponse("<script>alert('Email sent successfully.');window.location='/signin'</script>")

            except Login.DoesNotExist:
                return HttpResponse("<script>alert('Invalid email address.');window.location='/forgotpwd'</script>")
        else:
            return HttpResponse("<script>alert('Please enter a valid email address.');window.location='/forgotpwd'</script>")

    return redirect('/forgotpwd')
def addAdmin(request):
        Name = request.POST.get('username')
        EmailId = request.POST.get('email')
        Password = request.POST.get('password')
        ob = Login()
        ob.username = EmailId
        ob.password = Password
        ob.type = 'admin'
        ob.save()

        ob1 = Admin()
        ob1.Lid = ob
        ob1.Name = Name
        ob1.email = EmailId
        ob1.save()

        return HttpResponse("<script>alert('Inserted successfully');window.location='/signin'</script>")

def login(request):
    username = request.POST['username']
    password = request.POST['password']
    print(username,password,"6782292")
    try:
        print('scdef brgh')
        ob = Login.objects.get(username=username, password=password)
        print(ob,'12345678')
        if ob.type == "admin":
            request.session['alid'] = ob.id
            return HttpResponse("<script>window.location='/dashboard'</script>")
        elif ob.type == "doctor":
            request.session['dlid'] = ob.id
            return HttpResponse("<script>window.location='/doctordashboard'</script>")
        else:
            return HttpResponse("<script>alert('Invalid username or password');window.location='/signin'</script>")
    except Login.DoesNotExist:
        return HttpResponse("<script>alert('Invalid username password');window.location='/signin'</script>")
def header(request):
    return render(request,'components/header.html')
def header1(request):
    return render(request,'components/header1.html')
def viewTherapist(request):
    login_id = request.session.get('alid')
    therapist=Therapist.objects.filter(admin__Lid_id=login_id)
    context={
        'therapists':therapist
    }
    return render(request,'main/listTherapist.html',context)
def editpatientform(request,id):
    patient = Patient.objects.get(id=id)
    request.session['pid'] = patient.id
    print( request.session['pid'],'123456779')
    status=Status.objects.all()
    doctors = Doctor.objects.all()
    Therapies=Therapy.objects.all()
    context={
        'patient':patient,
        'doctors':doctors,
        'Therapies':Therapies,
        'date':str(patient.DOB),
        'status':status
    }
    return render(request,'main/editpatient.html',context)
def editpatient(request):
    name = request.POST['name']
    email = request.POST['email']
    dob = request.POST['dob']
    gender = request.POST['gender']
    phno = request.POST['phno']
    doctor = request.POST['doctor']
    status = request.POST['status']
    note = request.POST['note']
    ob1 = Patient.objects.get(id=request.session['pid'])
    status_instance = Status.objects.get(pk=status)
    if doctor:
                doctor_instance = Doctor.objects.get(pk=doctor)
                ob1.doctor = doctor_instance
    else:
         ob1.doctor = None
    if status:
                ob1.status = status_instance
    else:
         ob1.status = None
    ob1.name=name
    ob1.email=email
    ob1.DOB=dob
    ob1.gender=gender
    ob1.phone=phno
    # ob1.doctor=doctor_instance
    ob1.note=note
    ob1.save()
    if status:
            status_instance = Status.objects.get(pk=status)
            booking.objects.filter(patientid=ob1).update(status=status_instance)

    return HttpResponse("<script>alert('Updated successfully');window.location='/patient'</script>")
def deletePatient(request,id):
    ob = Patient.objects.get(id=id)
    ob.delete()
    book = request.GET.get('book')
    return HttpResponse("<script>alert('deleted successfully');window.location='/report'</script>")
def deleteTherapist(request, id):
    ob = Therapist.objects.get(id=id)
    ob.delete()
    return HttpResponse("<script>alert('Deleted successfully');window.location='/viewTherapist'</script>")
def deleteDoctor(request, id):
    ob = Doctor.objects.get(id=id)
    ob.delete()
    return HttpResponse("<script>alert('Deleted successfully');window.location='/viewDoctor'</script>")

def bookingview(request):
    search_query = request.GET.get('searchitem', '')
    per_page = int(request.GET.get('per_page', 10))
    login_id = request.session.get('alid')# Default to 10 if 'per_page' is not provided
    book_list = booking.objects.filter(admin__Lid_id=login_id).order_by('-id')

    if search_query:
        book_list = book_list.filter(patientid__name__icontains=search_query)

    paginator = Paginator(book_list, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request,'main/bookingview.html',context)
def deleteBooking(request,id):
    ob = booking.objects.get(id=id)
    ob.delete()
    return HttpResponse("<script>alert('deleted successfully');window.location='/patient'</script>")
def editbookingform(request,id):
    Booking = booking.objects.get(id=id)
    request.session['bid'] = Booking.id
    doctors = Doctor.objects.all()
    patient=Patient.objects.all()
    Therapies=Therapy.objects.all()
    treatment=Treatment.objects.all()
    context={
        'Booking':Booking,
        'doctors':doctors,
        'patient':patient,
        'date':str(Booking.reg_date),
        'time':str(Booking.reg_time),
        'treatment':treatment
    }
    return render(request,'main/editbooking.html',context)
def editbooking(request):
    doctor = request.POST['doctor']
    patient = request.POST['patient']
    regdate = request.POST['regdate']
    regtime = request.POST['regtime']
    treatment = request.POST['treatment']
    about = request.POST['about']
    ob1 = booking.objects.get(id=request.session['bid'])
    doctor_instance = Doctor.objects.get(pk=doctor)
    patient_instance = Patient.objects.get(pk=patient)
    treatment_instance = Treatment.objects.get(pk=treatment)
    ob1.Doctor=doctor_instance
    ob1.patientid=patient_instance
    ob1.reg_date=regdate
    ob1.reg_time=regtime
    ob1.about=about
    ob1.treatment=treatment_instance
    ob1.save()
    return HttpResponse("<script>alert('Updated successfully');window.location='/bookingview'</script>")
def schedules(request):
    doctors = Doctor.objects.all()
    patient=Patient.objects.all()
    Therapies=Therapy.objects.all()
    Therapists=Therapist.objects.all()
    context={
        'doctors':doctors,
        'patient':patient,
        'Therapies':Therapies,
        'Therapists':Therapists
    }
    return render(request,'main/scheduling.html',context)
def doctorprofile(request,id):
    doctor=Doctor.objects.get(id=id)
    department=Department.objects.all()
    request.session['docid'] = doctor.id
    context={
        'doctor':doctor,
        'department':department
    }
    return render(request,'main/doctorprofile.html',context)
def editdoctor(request):
    origin=request.POST['origin']
    name = request.POST['name']
    email = request.POST['email']
    gender = request.POST['gender']
    phno = request.POST['phno']
    place = request.POST['place']
    spaciality = request.POST['speciality']
    about = request.POST['about']
    image = request.FILES.get('image', None)
    ob1 = Doctor.objects.get(id=request.session['docid'])
    ob1.name=name
    ob1.email_id=email
    ob1.gender=gender
    ob1.contact_no=phno
    ob1.place=place
    ob1.Speciality=spaciality
    ob1.About=about
    if image:
            ob1.Image = image
    ob1.save()
    if origin == 'admin':
        return HttpResponse(f"<script>alert('Updated successfully');window.location='/doctorprofile/{ob1.id}'</script>")
    elif origin == 'doctor':
        return HttpResponse(f"<script>alert('Updated successfully');window.location='/docprofile'</script>")
def therapistprofile(request,id):
    therapist=Therapist.objects.get(id=id)
    request.session['tid'] = therapist.id
    Therapies=Therapy.objects.all()
    treatment=Treatment.objects.all()
    context={
        'therapist':therapist,
        'Therapies':Therapies,
        'treatment':treatment

    }
    return render(request,'main/therapistprofile.html',context)
@csrf_exempt
def edittherapist(request):
    name = request.POST['name']
    email = request.POST['email']
    gender = request.POST['gender']
    phno = request.POST['phno']
    place = request.POST['place']
    therapy = request.POST['speciality']
    about = request.POST['about']
    image = request.FILES.get('image', None)
    ob1 = Therapist.objects.get(id=request.session['tid'])
    Therapy_instance = Therapy.objects.get(pk=therapy)
    ob1.name=name
    ob1.email_id=email
    ob1.gender=gender
    ob1.contact_no=phno
    ob1.place=place
    ob1.specialization=Therapy_instance
    ob1.About=about
    if image:
            ob1.Image = image
    ob1.save()
    return HttpResponse(f"<script>alert('Updated successfully');window.location='/therapistprofile/{ob1.id}'</script>")

def addScheduling(request):
    login_id = request.session.get('alid')
    admin = Admin.objects.filter(Lid_id=login_id).first()

    patient = request.POST.get('patient', None)
    therapist = request.POST.get('therapist', None)
    apdate_str = request.POST.get('apdate', None)
    therapy = request.POST.get('therapy', None)
    note = request.POST.get('note', None)
    stime = request.POST.get('stime', None)
    etime = request.POST.get('etime', None)

    # Retrieve session_count and handle if it's null or empty
    session_count_str = request.POST.get('session', None)
    session_count = int(session_count_str) if session_count_str and session_count_str.isdigit() else None

    # Convert apdate to date object
    try:
        apdate = datetime.strptime(apdate_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid appointment date format. Please use YYYY-MM-DD.')
        return redirect('/schedules')

    obtherapy = Therapy.objects.get(id=therapy)
    obtherapist = Therapist.objects.get(id=therapist)
    obpatient = Patient.objects.get(id=patient)
    status, created = Status.objects.get_or_create(Status='Not Confirmed')

    # Check for conflicting schedules
    conflicting_schedules = Schedule.objects.filter(
        Therapist=obtherapist,
        Date=apdate,
        starting_Time__lt=etime,
        ending_Time__gt=stime
    )
    if conflicting_schedules.exists():
        conflicting_schedule_times = ", ".join(
            [f"{schedule.starting_Time.strftime('%H:%M')} to {schedule.ending_Time.strftime('%H:%M')}" for schedule in conflicting_schedules]
        )
        messages.error(request, f'This therapist is already scheduled at {conflicting_schedule_times}. Please select a different time or therapist.')
        return redirect('/schedules')
    try:
        latest_session = SessionStatus.objects.latest('id')
        session_group = latest_session.session_group + 1
    except SessionStatus.DoesNotExist:
        session_group = 1

    # Create the schedule for the first session
    # for session_num in range(1, session_count + 1):
    #     # Calculate the date for each session
    #     session_date = apdate + timedelta(days=(session_num - 1) * 7)
    for session_num in range(session_count):
        session_date = apdate + timedelta(days=session_num)
        # Create the schedule for each session
        ob1 = Schedule(
            admin=admin,
            patientid=obpatient,
            Therapist=obtherapist,
            Therapy=obtherapy,
            Note=note,
            Date=session_date,
            number_of_session=session_count,
            starting_Time=stime,
            ending_Time=etime,
            status=status
        )
        ob1.save()

        # Create session status record
        SessionStatus.objects.create(
            schedule=ob1,
            session_number=session_num,
            session_date=session_date,
            status=status.Status,  # Assuming status.Status is a string
            starting_Time=stime,
            ending_Time=etime,
            session_group=session_group
        )

    # Update patient status
    obpatient.status = status
    obpatient.save()

    return HttpResponse("<script>alert('Inserted successfully');window.location='/book'</script>")

# def addScheduling(request):
#     login_id = request.session.get('alid')
#     admin = Admin.objects.filter(Lid_id=login_id).first()

#     patient_id = request.POST.get('patient', None)
#     therapist_id = request.POST.get('therapist', None)
#     apdate_str = request.POST.get('apdate', None)
#     therapy_id = request.POST.get('therapy', None)
#     note = request.POST.get('note', None)
#     stime = request.POST.get('stime', None)
#     etime = request.POST.get('etime', None)
#     session_count_str = request.POST.get('session', None)
#     session_count = int(session_count_str) if session_count_str and session_count_str.isdigit() else None

#     # Convert apdate to date object
#     try:
#         apdate = datetime.strptime(apdate_str, '%Y-%m-%d').date()
#     except ValueError:
#         messages.error(request, 'Invalid appointment date format. Please use YYYY-MM-DD.')
#         return redirect('/schedules')

#     obtherapy = Therapy.objects.get(id=therapy_id)
#     obtherapist = Therapist.objects.get(id=therapist_id)
#     obpatient = Patient.objects.get(id=patient_id)
#     status, created = Status.objects.get_or_create(Status='Pending')

#     # Check for conflicting schedules
#     conflicting_schedules = Schedule.objects.filter(
#         Therapist=obtherapist,
#         Date=apdate,
#         starting_Time__lt=etime,
#         ending_Time__gt=stime
#     )
#     if conflicting_schedules.exists():
#         conflicting_schedule_times = ", ".join(
#             [f"{schedule.starting_Time.strftime('%H:%M')} to {schedule.ending_Time.strftime('%H:%M')}" for schedule in conflicting_schedules]
#         )
#         messages.error(request, f'This therapist is already scheduled at {conflicting_schedule_times}. Please select a different time or therapist.')
#         return redirect('/schedules')

#     # Check for existing sessions for the same patient and therapist
#     existing_sessions = SessionStatus.objects.filter(
#         schedule__Therapist=obtherapist,
#         schedule__patientid=obpatient,
#         schedule__Date=apdate
#     )

#     if existing_sessions.exists():
#         # Update existing sessions with new date
#         for session in existing_sessions:
#             session.session_date = apdate
#             session.save()
#     else:
#         # Create the schedule
#         ob1 = Schedule(
#             admin=admin,
#             patientid=obpatient,
#             Therapist=obtherapist,
#             Therapy=obtherapy,
#             Note=note,
#             Date=apdate,
#             number_of_session=session_count if session_count is not None else 0,
#             starting_Time=stime,
#             ending_Time=etime,
#             status=status
#         )
#         ob1.save()

#         # Create session status records if session_count is provided
        # if session_count:
        #    existing_sessions = SessionStatus.objects.filter(schedule=ob1)
        #    for session_num in range(1, session_count + 1):
        #     if session_num <= existing_sessions.count():
        #         # Update existing sessions with the new date if it is the first session
        #         if session_num == 1 and existing_sessions.filter(session_number=session_num).exists():
        #             session = existing_sessions.get(session_number=session_num)
        #             session.session_date = apdate
        #             session.save()
        #     else:
        #         # Create new sessions if not already existing
        #         session_date = apdate if session_num == 1 else None
        #         SessionStatus.objects.create(
        #             schedule=ob1,
        #             session_number=session_num,
        #             session_date=session_date,
        #             status='pending'
        #         )


#     # Update patient status
#     obpatient.status = status
#     obpatient.save()

#     return HttpResponse("<script>alert('Inserted successfully');window.location='/book'</script>")
def listschedules(request):
    search_query = request.GET.get('searchitem', '')
    per_page = int(request.GET.get('per_page', 10))
    login_id = request.session.get('alid')
    schedule_list = Schedule.objects.filter(admin__Lid_id=login_id).order_by('-id')
    patient_id = request.GET.get('patient_id')
    if patient_id:
        schedule_list = schedule_list.filter(id=patient_id)
    # if search_query:
    #     schedule_list = schedule_list.filter(patientid__name__icontains=search_query)

    paginator = Paginator(schedule_list, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'main/listschedules.html', context)
def view_sessions(request, schedule_id):
    # Retrieve the schedule by its ID
    schedule = Schedule.objects.get(id=schedule_id)

    # Get patient and therapist from the schedule
    patient = schedule.patientid
    therapist = schedule.Therapist
    therapy=schedule.Therapy
    # Filter sessions based on patient and therapist
    sessions = SessionStatus.objects.filter(
        schedule__patientid=patient,
        schedule__Therapist=therapist
    ).order_by('-session_group', 'session_number')

    # Pass the sessions to the template
    return render(request, 'main/sessions.html', {'sessions': sessions, 'patient': patient, 'therapist': therapist,'therapy':therapy})



# def view_sessions(request, schedule_id):
#     print('23')
#     schedule = get_object_or_404(Schedule, id=schedule_id)
#     sessions = SessionStatus.objects.filter(schedule=schedule)
#     print(sessions,'lolpolpolpolp')
#     # Create sessions if they do not exist
#     if not sessions.exists():
#         for i in range(1, int(schedule.number_of_session) + 1):
#             SessionStatus.objects.create(schedule=schedule, session_number=i)
#         sessions = SessionStatus.objects.filter(schedule=schedule)

#     context = {
#         'schedule': schedule,
#         'sessions': sessions,
#     }
#     return render(request, 'main/sessions.html', context)

def toggle_session_status(request, session_id):
    print('hi ')
    session = get_object_or_404(SessionStatus, id=session_id)
    session.status = 'attended' if session.status == 'Not Confirmed' else 'Not Confirmed'
    status, created = Status.objects.get_or_create(Status='Attended')
    session.schedule.status=status
    session.schedule.save()
    # Booking = booking.objects.filter(patientid=session.schedule.patientid).first()  # Get related booking
    # if Booking:
    #     Booking.status = status
    #     Booking.save()
    session.save()
    return redirect('view_sessions', schedule_id=session.schedule.id)

def editschedulingform(request,id):
    schedule=Schedule.objects.get(id=id)
    patient=Patient.objects.all()
    therapist=Therapist.objects.all()
    request.session['schid'] = schedule.id
    status=Status.objects.all()
    context={
        'schedule':schedule,
        'date':str(schedule.Date),
        'stime':str(schedule.starting_Time),
        'etime':str(schedule.ending_Time),
        'patient':patient,
        'therapist':therapist,
        'status':status
    }
    return render(request,'main/editscheduling.html',context)
def editschedule(request):
    patient=request.POST.get('patient', None)
    therapist=request.POST.get('therapist', None)
    apdate=request.POST.get('apdate', None)
    therapy=request.POST.get('therapy', None)
    session=request.POST.get('session', None)
    note=request.POST.get('note', None)
    stime=request.POST.get('stime')
    etime=request.POST.get('etime')
    status = request.POST['status']
    ob1 = Schedule.objects.get(id=request.session['schid'])
    obtherapy = Therapy.objects.get(pk=therapy)
    obtherapist = Therapist.objects.get(pk=therapist)
    obpatient = Patient.objects.get(pk=patient)
    status_instance = Status.objects.get(pk=status)
    ob1.patientid=obpatient
    ob1.Therapist=obtherapist
    ob1.Therapy=obtherapy
    ob1.Note=note
    ob1.Date=apdate
    ob1.number_of_session=session
    ob1.starting_Time=stime
    ob1.ending_Time=etime
    ob1.status=status_instance
    ob1.save()
    sessions_to_update = SessionStatus.objects.filter(schedule=ob1)

    for session in sessions_to_update:
        session.session_date = apdate
        session.starting_Time = stime
        session.ending_Time = etime
        session.save()
    return HttpResponse("<script>alert('Updated successfully');window.location='/listschedules'</script>")
def deleteSchedule(request,id):
    ob = Schedule.objects.get(id=id)
    ob.delete()
    return HttpResponse("<script>alert('deleted successfully');window.location='/listschedules'</script>")
def viewDoctor(request):
    login_id = request.session.get('alid')
    doctor=Doctor.objects.filter(admin__Lid_id=login_id)
    context={
        'doctors':doctor
    }
    return render(request,'main/listdoctor.html',context)
def searchpatient(request):
   if request.method == 'POST':
        search_term = request.POST.get('searchitem', '')
        task = Patient.objects.all()
        results = task.filter(
            Q(name__icontains=search_term) |
            Q(number_of_session__icontains=search_term) |
            Q(medicines__icontains=search_term) |
            Q(note__icontains=search_term) |
            Q(blood_gp__icontains=search_term) |
            Q(address__icontains=search_term) |
            Q(email__icontains=search_term) |
            Q(patient_id__icontains=search_term) |
            Q(gender__icontains=search_term) |
            Q(DOB__icontains=search_term) |
            Q(age__icontains=search_term) |
            Q(status__Status__icontains=search_term) |
            Q(doctor__name__icontains=search_term) |
            Q(therapy__Therapy__icontains=search_term)

        )
        page_number = request.GET.get('page', 1)  # Get the page number from the request
        paginator = Paginator(results, 10)  # Show 10 results per page
        page_obj = paginator.get_page(page_number)
        context = {
            'page_obj': page_obj,

        }
        return render(request, 'main/patient.html', context)
def searchbooking(request):
    if request.method == 'POST':
        search_term = request.POST.get('searchitem', '')
        task = booking.objects.all()
        results = task.filter(
            Q(reg_time__icontains=search_term) |
            Q(reg_date__icontains=search_term) |
            Q(about__icontains=search_term) |
            Q(Doctor__name__icontains=search_term) |
            Q(treatment__Treatment__icontains=search_term) |
            Q(patientid__name__icontains=search_term) |
            Q(patientid__patient_id__icontains=search_term)

        )
        page_number = request.GET.get('page', 1)  # Get the page number from the request
        paginator = Paginator(results, 10)  # Show 10 results per page
        page_obj = paginator.get_page(page_number)
        context = {
            'page_obj': page_obj,

        }

        return render(request,'main/bookingview.html', context)
def searchschedule(request):
    if request.method == 'POST':
        # Handle the search functionality
        search_term = request.POST.get('searchitem', '')
        results = Schedule.objects.filter(
            Q(ending_Time__icontains=search_term) |
            Q(starting_Time__icontains=search_term) |
            Q(number_of_session__icontains=search_term) |
            Q(Date__icontains=search_term) |
            Q(Note__icontains=search_term) |
            Q(patientid__name__icontains=search_term) |
            Q(patientid__patient_id__icontains=search_term) |
            Q(Therapist__name__icontains=search_term) |
            Q(Therapy__Therapy__icontains=search_term)
        )
    else:
        # Handle GET requests by using the search term from GET parameters
        search_term = request.GET.get('searchitem', '')
        results = Schedule.objects.filter(
            Q(ending_Time__icontains=search_term) |
            Q(starting_Time__icontains=search_term) |
            Q(number_of_session__icontains=search_term) |
            Q(Date__icontains=search_term) |
            Q(Note__icontains=search_term) |
            Q(patientid__name__icontains=search_term) |
            Q(patientid__patient_id__icontains=search_term) |
            Q(Therapist__name__icontains=search_term) |
            Q(Therapy__Therapy__icontains=search_term)
        )

    page_number = request.GET.get('page', 1)  # Get the page number from the request
    paginator = Paginator(results, 10)  # Show 10 results per page
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'main/listschedules.html', context)

# def searchschedule(request):
#     if request.method == 'POST':
#         search_term = request.POST.get('searchitem', '')
#         task = Schedule.objects.all()
#         results = task.filter(
#             Q(ending_Time__icontains=search_term) |
#             Q(starting_Time__icontains=search_term) |
#             Q(number_of_session__icontains=search_term) |
#             Q(Date__icontains=search_term) |
#             Q(Note__icontains=search_term) |
#             Q(patientid__name__icontains=search_term) |
#             Q(patientid__patient_id__icontains=search_term) |
#             Q(Therapist__name__icontains=search_term) |
#             Q(Therapy__Therapy__icontains=search_term)

#         )

#         page_number = request.GET.get('page', 1)  # Get the page number from the request
#         paginator = Paginator(results, 10)  # Show 10 results per page
#         page_obj = paginator.get_page(page_number)
#         context = {
#             'page_obj': page_obj,

#         }
#         return render(request,'main/listschedules.html', context)
def patientprofile(request,id):
    patient = Patient.objects.get(id=id)
    request.session['proid'] = patient.id
    context={
        'patient':patient,
    }
    return render(request,'main/patientprofile.html',context)
def random_color():
    return "#{:06x}".format(random.randint(0, 0xFFFFFF))
def calendar(request):
    therapist_id = request.GET.get('therapist')
    patient_id = request.GET.get('patient')
    login_id = request.session.get('alid')
    tasks =  Schedule.objects.filter(admin__Lid_id=login_id)
    if therapist_id:
        print('12345')
        tasks = tasks.filter(Therapist_id=therapist_id)
    if patient_id:
        print('hi')
        tasks = tasks.filter(patientid_id=patient_id)
    therapist=Therapist.objects.all()
    patient=Patient.objects.all()
    events = []

    for task in tasks:
        event = {
            'title': f"{task.patientid.name} with {task.Therapist.name}",
            'start': str(task.Date) + 'T' + str(task.starting_Time),
            'end': str(task.Date) + 'T' + str(task.ending_Time),
            'color': random_color(),
        }
        events.append(event)
    context = {
        'events': events,
        'therapist':therapist,
        'patient':patient

    }
    return render(request, 'main/calendar.html', context)
def book(request):
    login_id = request.session.get('alid')
    tasks = Schedule.objects.filter(admin__Lid_id=login_id)
    therapist=Therapist.objects.filter(admin__Lid_id=login_id)
    patient=Patient.objects.filter(admin__Lid_id=login_id)
    selected_patient = request.GET.get('patient')
    selected_therapist = request.GET.get('therapist')
    date_filter = request.GET.get('date_filter')
    custom_date = request.GET.get('custom_date')
    status=Status.objects.all()
    if selected_patient:
        tasks = tasks.filter(patientid__id=selected_patient)
    if selected_therapist:
        print('yes yo ')
        tasks = tasks.filter(Therapist__id=selected_therapist)
    if date_filter:
        today = datetime.today().date()
        if date_filter == 'today':
            tasks = tasks.filter(Date=today)
        elif date_filter == 'yesterday':
            tasks = tasks.filter(Date=today - timedelta(days=1))
        elif date_filter == 'custom' and custom_date:
            from_date_parsed = parse_date(custom_date)
            tasks = tasks.filter(Date=from_date_parsed)

    events = []

    therapist_dict = {t.id: index + 1 for index, t in enumerate(therapist)}

    for task in tasks:
        patient_status = task.status
        event = {
            'title': f"{task.patientid.name} with {task.Therapist.name}",
            'start': str(task.Date) + 'T' + str(task.starting_Time),
            'end': str(task.Date) + 'T' + str(task.ending_Time),
            'color':task.status.color if patient_status else  random_color(),
            'column': therapist_dict[task.Therapist.id],
            'patient': task.patientid.name,
            'id': task.patientid.patient_id,
            'therapy': task.Therapy.Therapy,
            'note': task.Note,
            'patientId':task.id
        }
        events.append(event)
        print('yaya',events)
    current_date = datetime.now()
    formatted_date = current_date.strftime('%B %d, %Y')
    weekday = current_date.strftime('%A')
    context = {
        'events': events,
        'therapist': therapist,
        'patient': patient,
        'current_date': formatted_date,
        'current_weekday': weekday,
        'tasks':tasks,
        'status':status
    }
    return render(request, 'main/book.html', context)
def update_schedule(request):
    print('qe')
    date = request.GET.get('date')
    selected_therapist = request.GET.get('therapist')
    login_id = request.session.get('alid')
    tasks = Schedule.objects.filter(admin__Lid_id=login_id)
    if date:
        date = parse_date(date)
        tasks = tasks.filter(Date=date)
        if selected_therapist:
           print('yes bee ')
           tasks = tasks.filter(Therapist__id=selected_therapist)

        events = []
        therapist = Therapist.objects.filter(admin__Lid_id=login_id)
        therapist_dict = {t.id: index + 1 for index, t in enumerate(therapist)}

        for task in tasks:
            patient_status = task.status
            event = {
                'title': f"{task.patientid.name} with {task.Therapist.name}",
                'start': str(task.Date) + 'T' + str(task.starting_Time),
                'end': str(task.Date) + 'T' + str(task.ending_Time),
                'color':task.status.color if patient_status else  random_color(),
                'column': therapist_dict[task.Therapist.id],
                'patient': task.patientid.name,
                'id': task.patientid.patient_id,
                'therapy': task.Therapy.Therapy,
                'note': task.Note,
                'patientId':task.id
            }
            events.append(event)
            print('cvbn',events)
        return JsonResponse({'events': events})
    return JsonResponse({'events': []})
def therapy(request):
    query = request.GET.get('search', '')
    login_id = request.session.get('alid')
    per_page = int(request.GET.get('per_page', 10))# Get the search query from the GET parameters
    if request.method == 'POST':
        therapy_name = request.POST.get('therapy')
        login_id = request.session.get('alid')
        admin = Admin.objects.filter(Lid_id=login_id).first()
        if therapy_name:
            new_therapy = Therapy(Therapy=therapy_name)
            new_therapy.admin=admin
            new_therapy.save()
            return redirect('therapy')

    if query:
        therapies = Therapy.objects.filter(Therapy__icontains=query)
    else:
        therapies = Therapy.objects.filter(admin__Lid_id=login_id)
    paginator = Paginator(therapies, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    return render(request, 'main/therapy.html', {'page_obj': page_obj, 'search_query': query})
def delete_therapy(request, id):
    therapy = Therapy.objects.get(id=id)
    therapy.delete()
    return HttpResponse("<script>alert('deleted successfully');window.location='/therapy'</script>")
def edittherapy(request ,id):
    therapy = get_object_or_404(Therapy, id=id)
    if request.method == 'POST':
        therapy_name = request.POST.get('therapy')
        therapy.Therapy = therapy_name
        therapy.save()
        return redirect('therapy')
    return render(request,'main/edittherapy.html',{'therapy': therapy})
@csrf_exempt
def update_profile_picture(request):
    if request.method == 'POST' and request.FILES.get('image'):
        doctor_id = request.POST.get('doctor_id')  # Adjust this as per your requirement
        doctor = get_object_or_404(Doctor, id=doctor_id)
        doctor.Image = request.FILES['image']
        doctor.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)
# def doctorpatient(request):
#     login_id = request.session.get('dlid')
#     doctor = Doctor.objects.filter(Lid_id=login_id).first()
#     search_query = request.GET.get('searchitem', '')
#     per_page = int(request.GET.get('per_page', 10))  # Default to 10 if 'per_page' is not provided
#     # patient_list = booking.objects.all()
#     patient_list = booking.objects.filter(Doctor=doctor,patientid__status__Status__in=['Confirmed', 'Registered']).order_by('-id')
#     if search_query:
#         patient_list = patient_list.filter(
#             Q(patientid__name__icontains=search_query) |  # Assuming 'name' field in patientid
#             Q(patientid__patient_id__icontains=search_query)  # Assuming 'patient_id' field in patientid
#         )

#     paginator = Paginator(patient_list, per_page)
#     page_number = request.GET.get('page')  # Get the page number from the URL
#     page_obj = paginator.get_page(page_number)

#     context = {
#         'page_obj': page_obj,
#         'search_query': search_query,
#     }
#     return render(request,'main/doctorpatient.html',context)
def doctorpatient(request):
    login_id = request.session.get('dlid')
    doctor = Doctor.objects.filter(Lid_id=login_id).first()
    search_query = request.GET.get('searchitem', '')
    per_page = int(request.GET.get('per_page', 10))  # Default to 10 if 'per_page' is not provided

    # Filter bookings for the logged-in doctor with specific statuses
    patient_list = booking.objects.filter(
        Doctor=doctor,
        patientid__status__Status__in=['Confirmed', 'Registered']
    ).order_by('-id')

    # Apply filtering based on search query
    if search_query:
        patient_list = patient_list.filter(
            Q(patientid__name__icontains=search_query) |
            Q(patientid__patient_id__icontains=search_query) |
            Q(patientid__status__Status__icontains=search_query) |
            Q(patientid__email__icontains=search_query)
        )

    # Implement pagination
    paginator = Paginator(patient_list, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    # Pass context to the template
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'main/doctorpatient.html', context)
def searchdoctorpatient(request):
    # Assuming the search form submits via GET, not POST
    login_id = request.session.get('dlid')
    doctor = Doctor.objects.filter(Lid_id=login_id).first()

    # Ensure the search query is pulled from GET request
    search_query = request.GET.get('searchitem', '')

    # Fetch patients linked to the logged-in doctor with specific statuses
    patient_list = booking.objects.filter(
        Doctor=doctor,
        patientid__status__Status__in=['Confirmed', 'Registered']
    ).order_by('-id')

    # Apply search filters
    if search_query:
        patient_list = patient_list.filter(
            Q(patientid__name__icontains=search_term) |
            Q(patientid__patient_id__icontains=search_query) |
            Q(patientid__medicines__icontains=search_query) |
            Q(patientid__note__icontains=search_query) |
            Q(patientid__blood_gp__icontains=search_query) |
            Q(patientid__address__icontains=search_query) |
            Q(patientid__email__icontains=search_query) |
            Q(patientid__age__icontains=search_query) |
            Q(patientid__doctor__name__icontains=search_query) |
            Q(patientid__therapy__Therapy__icontains=search_query)
        )

    # Set up pagination
    page_number = request.GET.get('page', 1)  # Default to the first page if not provided
    paginator = Paginator(patient_list, 10)  # Show 10 results per page
    page_obj = paginator.get_page(page_number)

    # Pass results to the template
    context = {
        'page_obj': page_obj,
        'search_query': search_query,  # Keep the search query in the context for use in the template
    }
    return render(request, 'main/doctorpatient.html', context)
def editpatientreportform(request,id):
    patient = booking.objects.get(id=id)
    request.session['pnid'] = patient.patientid.id


    doctors = Doctor.objects.all()
    Therapies=Therapy.objects.all()
    context={
        'patient':patient.patientid,
        'doctors':doctors,
        'Therapies':Therapies,
        'date':str(patient.patientid.DOB),
    }
    return render(request,'main/patientreport.html',context)
def addpatientreportform(request):
        blood_gp = request.POST.get('bloodgp')
        therapy = request.POST.get('therapy')
        number_of_session = request.POST.get('number_of_session')
        medicines = request.POST.get('medicine')
        Digestion = request.POST.get('Digestion')
        Sleep = request.POST.get('Sleep')
        Allergies = request.POST.get('Allergies')
        Menstrualhistory = request.POST.get('Menstrualhistory')
        Presentingcomplaints = request.POST.get('Presentingcomplaints')
        Historyofpresentingcomplaints = request.POST.get('History')
        PastMedicalandsurgicalhistory = request.POST.get('PastMedical')
        Regularmedications = request.POST.get('Regularmedications')
        Amanirama = request.POST.get('Amanirama')
        Doshapredominence = request.POST.get('Dosha')
        Dhathupredominence = request.POST.get('Dhathu')
        Srothusinvolved = request.POST.get('Srothus')
        Treatment = request.POST.get('Treatment')
        Proposedtreatmentplan = request.POST.get('Proposedtreatmentplan')
        Followup = request.POST.get('Followup')
        ob1 = Patient.objects.get(id=request.session['pnid'])
        # Therapy_instance = Therapy.objects.get(pk=therapy)
        ob1.blood_gp=blood_gp
        if therapy:
            Therapy_instance = Therapy.objects.get(pk=therapy)
            ob1.therapy = Therapy_instance
        else:
            ob1.therapy = None
        # ob1.therapy=Therapy_instance
        ob1.medicines=medicines
        ob1.number_of_session=number_of_session
        ob1.Digestion=Digestion
        ob1.Sleep=Sleep
        ob1.Allergies=Allergies
        ob1.Menstrualhistory=Menstrualhistory
        ob1.Presentingcomplaints=Presentingcomplaints
        ob1.Historyofpresentingcomplaints=Historyofpresentingcomplaints
        ob1.PastMedicalandsurgicalhistory=PastMedicalandsurgicalhistory
        ob1.Regularmedications=Regularmedications
        ob1.Amanirama=Amanirama
        ob1.Doshapredominence=Doshapredominence
        ob1.Dhathupredominence=Dhathupredominence
        ob1.Srothusinvolved=Srothusinvolved
        ob1.Treatment=Treatment
        ob1.Proposedtreatmentplan=Proposedtreatmentplan
        ob1.Followup=Followup
        booking_status, created = Status.objects.get_or_create(Status='Attended')
        ob1.status = booking_status
        ob1.save()
        return HttpResponse("<script>alert('Updated successfully');window.location='/consultedpatient'</script>")
def casesheets(request,id):
    case=booking.objects.get(id=id)
    context={
        'case':case.patientid
    }
    return render(request,'main/casesheet.html',context)
def case(request,id):
    case=Patient.objects.get(id=id)
    context={
        'case':case
    }
    return render(request,'main/case.html',context)
def consultedpatient(request):
    login_id = request.session.get('dlid')
    doctor = Doctor.objects.filter(Lid_id=login_id).first()
    search_query = request.GET.get('searchitem', '')
    per_page = int(request.GET.get('per_page', 10))  # Default to 10 if 'per_page' is not provided
    # patient_list = booking.objects.filter(patientid__status="Consulted").order_by('-id')
    patient_list = booking.objects.filter(
    Doctor=doctor,
    patientid__status__Status__in=[ "Not Confirmed", "Attended"]
).order_by('-id')
    if search_query:
        patient_list = patient_list.filter(
            Q(patientid__name__icontains=search_query) |
            Q(patientid__patient_id__icontains=search_query) |
            Q(patientid__status__Status__icontains=search_query) |
            Q(patientid__email__icontains=search_query)
        )
    paginator = Paginator(patient_list, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request,'main/consultedpatient.html',context)
def report(request):
    login_id = request.session.get('alid')
    tasks = booking.objects.filter(admin__Lid_id=login_id)
    doctors = Doctor.objects.filter(admin__Lid_id=login_id)
    patients = Patient.objects.filter(admin__Lid_id=login_id)
    selected_patient = request.GET.get('patient')
    selected_doctor = request.GET.get('doctor')
    date_filter = request.GET.get('date_filter')
    custom_date = request.GET.get('custom_date')
    status=Status.objects.all()
    if selected_patient:
        tasks = tasks.filter(patientid__id=selected_patient)
    if selected_doctor:
        tasks = tasks.filter(Doctor__id=selected_doctor)
    if date_filter:
        today = datetime.today().date()
        if date_filter == 'today':
            tasks = tasks.filter(reg_date=today)
        elif date_filter == 'yesterday':
            tasks = tasks.filter(reg_date=today - timedelta(days=1))
        elif date_filter == 'custom' and custom_date:
            from_date_parsed = parse_date(custom_date)
            tasks = tasks.filter(reg_date=from_date_parsed)

    events = []

    doctor_dict = {d.id: index + 1 for index, d in enumerate(doctors)}

    for task in tasks:
        patient_status = task.patientid.status
        event = {
            'title': f"{task.patientid.name} with {task.Doctor.name}",
            'start': str(task.reg_date) + 'T' + str(task.reg_time),
            'end': str(task.reg_date) + 'T' + str(task.reg_time),
            'color':task.patientid.status.color if patient_status else  random_color(),
            'column': doctor_dict[task.Doctor.id],
            # 'color': random_color(),
            'patient': task.patientid.name,
            'about': task.treatment.Treatment if task.treatment else '',
            'phno':task.patientid.phone,
            'patientId':task.patientid.id
        }
        events.append(event)

    current_date = datetime.now()
    formatted_date = current_date.strftime('%B %d, %Y')
    weekday = current_date.strftime('%A')
    context = {
        'events': events,
        'doctors': doctors,
        'patients': patients,
        'current_date': formatted_date,
        'current_weekday': weekday,
        'tasks': tasks,
        'status':status
    }
    return render(request, 'main/report.html', context)

def update_schedules(request):
    date = request.GET.get('date')
    selected_doctor = request.GET.get('doctor')
    login_id = request.session.get('alid')
    tasks = booking.objects.filter(admin__Lid_id=login_id)
    if date:
        date = parse_date(date)
        tasks = tasks.filter(reg_date=date)
        if selected_doctor:
            tasks = tasks.filter(Doctor__id=selected_doctor)

        events = []
        doctors = Doctor.objects.filter(admin__Lid_id=login_id)
        doctor_dict = {d.id: index + 1 for index, d in enumerate(doctors)}

        for task in tasks:
            patient_status = task.patientid.status
            event = {
                'title': f"{task.patientid.name} with {task.Doctor.name}",
                'start': str(task.reg_date) + 'T' + str(task.reg_time),
                'end': str(task.reg_date) + 'T' + str(task.reg_time),
                'color':task.patientid.status.color if patient_status else  random_color(),
                # 'color': random_color(),
                'column': doctor_dict[task.Doctor.id],
                'patient': task.patientid.name,
                'about': task.treatment.Treatment if task.treatment else '',
                'phno':task.patientid.phone,
                'patientId':task.patientid.id


            }
            events.append(event)
            print(events)
        return JsonResponse({'events': events})
    return JsonResponse({'events': []})
def department(request):
    query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 10)

    if request.method == 'POST':
        department_name = request.POST.get('department')
        if department_name:
            new_department = Department(Department=department_name)
            new_department.save()
        return redirect('department')
    if query:
        departments = Department.objects.filter(Department__icontains=query)
    else:
        departments = Department.objects.all()

    paginator = Paginator(departments, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'main/department.html', {
        'departments': page_obj,
        'page_obj': page_obj,
        'search_query': query,

    })

def delete_department(request, id):
    department = get_object_or_404(Department, id=id)
    department.delete()
    return redirect('department')

def edit_department(request, id):
    department = get_object_or_404(Department, id=id)
    if request.method == 'POST':
        department_name = request.POST.get('department')
        department.Department = department_name
        department.save()
        return redirect('department')

    return render(request, 'main/edit_department.html', {'department': department})
def treatment(request):
    query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 10)

    if request.method == 'POST':
        treatment_name = request.POST.get('treatment')
        department_name = request.POST.get('department')
        obdepartment = Department.objects.get(id=department_name)
        ob1=Treatment()
        ob1.Treatment=treatment_name
        ob1.Department=obdepartment
        ob1.save()
        return redirect('treatment')

    if query:
        treatments = Treatment.objects.filter(
            Q(Treatment__icontains=query) |
            Q(Department__Department__icontains=query)
        )
    else:
        treatments = Treatment.objects.all()

    paginator = Paginator(treatments, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    department=Department.objects.all()
    return render(request, 'main/treatment.html', {
        'treatments': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'department':department
    })

def delete_treatment(request, id):
    treatment = get_object_or_404(Treatment, id=id)
    treatment.delete()
    return redirect('treatment')
def edit_treatment(request, id):
    treatment = get_object_or_404(Treatment, id=id)
    if request.method == 'POST':
        treatment_name = request.POST.get('treatment')
        treatment.Treatment = treatment_name
        treatment.save()
        return redirect('treatment')

    return render(request, 'main/edit_treatment.html', {'treatment': treatment})

def status(request):
    if request.method == 'POST':
        status_id = request.POST.get('status_id')
        color = request.POST.get('color')
        new_status = request.POST.get('new_status')

        if new_status:
            # Create a new Status object
            ob = Status(Status=new_status, color=color)
            ob.save()
            # Redirect or handle success
            return redirect('/report')  # Replace '/success' with your success URL

        if status_id and color:
            # Update existing Status object
            try:
                ob = Status.objects.get(pk=status_id)
                ob.color = color
                ob.save()
                return redirect('/report')  # Replace '/success' with your success URL
            except Status.DoesNotExist:
                return render(request, 'report.html', {'error': 'Status not found.'})

        return render(request, 'report.html', {'error': 'Please select or enter a status and choose a color.'})

    return render(request, 'report.html')

def statuss(request):
    if request.method == 'POST':
        status_id = request.POST.get('status_id')
        color = request.POST.get('color')
        new_status = request.POST.get('new_status')

        if new_status:
            # Create a new Status object
            ob = Status(Status=new_status, color=color)
            ob.save()
            # Redirect or handle success
            return redirect('/book')  # Replace '/success' with your success URL

        if status_id and color:
            # Update existing Status object
            try:
                ob = Status.objects.get(pk=status_id)
                ob.color = color
                ob.save()
                return redirect('/book')  # Replace '/success' with your success URL
            except Status.DoesNotExist:
                return render(request, 'bokk.html', {'error': 'Status not found.'})

        return render(request, 'book.html', {'error': 'Please select or enter a status and choose a color.'})

    return render(request, 'book.html')

def deletestatus(request, id):
    if request.method == 'POST':
        ob = get_object_or_404(Status, id=id)
        ob.delete()
        return redirect('/report')
    return redirect('/report')

def deletestatuss(request, id):
    if request.method == 'POST':
        ob = get_object_or_404(Status, id=id)
        ob.delete()
        return redirect('/book')
    return redirect('/book')
def patient_profile(request, patient_id):
    print('hi me')
    patient = get_object_or_404(Patient, id=patient_id)
    context = {
        'patient': patient
    }
    return render(request, 'main/patient.html', context)
def adminprofile(request):
    login_id = request.session.get('alid')
    admin = Admin.objects.filter(Lid_id=login_id)
    print(admin)
    context={
        'admin':admin
    }
    return render(request,'main/adminprofile.html',context)
def changepwd(request):
    return render(request,'main/chnagepwd.html')
def editadmin(request):
    name = request.POST['name']
    email = request.POST['email']
    phno = request.POST['phno']
    image = request.FILES.get('image', None)
    login_id = request.session.get('alid')
    login_instance = Login.objects.get(id=login_id)
    ob1 = Admin.objects.get(Lid=login_instance)
    ob1.Name=name
    ob1.email=email
    ob1.phone=phno
    if image:
            ob1.Image = image
    ob1.save()
    return HttpResponse(f"<script>alert('Updated successfully');window.location='/adminprofile'</script>")
def passwordchange(request):
    cpwd=request.POST['currentpwd']
    npwd=request.POST['newpwd']
    cnpwd=request.POST['confirmnewpwd']
    origin=request.POST['origin']
    try:
        if origin == 'admin':
          ob=Login.objects.get(password=cpwd,id=request.session['alid'])
        elif origin == 'doctor':
          ob=Login.objects.get(password=cpwd,id=request.session['dlid'])
        if ob is not None:
            if npwd == cnpwd:
                ob.password=npwd
                ob.save()
                if origin == 'admin':
                   return HttpResponse("<script> alert('password changed  successfully');window.location='/dashboard'</script>")
                elif origin == 'doctor':
                   return HttpResponse("<script> alert('password changed  successfully');window.location='/doctorpatient'</script>")
            else:
                if origin == 'admin':
                    return HttpResponse("<script> alert(' password mismatch');window.location='/changepwd'</script>")
                elif origin == 'doctor':
                    return HttpResponse("<script> alert(' password mismatch');window.location='/changepwd'</script>")

    except:
        if origin == 'admin':
            return HttpResponse("<script> alert('incorrect password');window.location='/changepwd'</script>")
        elif origin == 'doctor':
                    return HttpResponse("<script> alert(' password mismatch');window.location='/changepwd'</script>")


def doctorchangepwd(request):
    return render(request,'main/doctorchangepwd.html')
def docprofile(request):
    login_id = request.session.get('dlid')
    doctor = Doctor.objects.filter(Lid_id=login_id)
    print(doctor,'nji')
    context={
        'doctors':doctor
    }
    return render(request,'main/doctorprofiledashboard.html',context)
def doctordashboard(request):
    login_id = request.session.get('dlid')
    per_page = int(request.GET.get('per_page', 10))
    # Get the logged-in doctor based on the login ID
    doctor = get_object_or_404(Doctor, Lid_id=login_id)

    # Filter bookings for the logged-in doctor with specific patient statuses
    bookings = booking.objects.filter(
        Doctor=doctor,
        patientid__status__Status__in=['Confirmed', 'Registered', '	Not Confirmed','Attended']
    ).order_by('-id')
    total_patients = bookings.count()
    today = timezone.now().date()
    today_created_count = bookings.filter(created_at__date=today).count()
    consultedpatient = booking.objects.filter(
        Doctor=doctor,
        patientid__status__Status='Attended'
    )

    # Count the number of consulted patients
    consultedpatientcount = consultedpatient.count()
    paginator = Paginator(bookings, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'total_patients':total_patients,
        'today_created_count':today_created_count,
        'consultedpatientcount':consultedpatientcount
    }
    return render(request,'main/doctordashboard.html',context)

def history(request,id):
    patient = get_object_or_404(Patient, id=id)

    # Fetch all schedules and bookings for the patient
    schedules = Schedule.objects.filter(patientid=patient).order_by('-Date', '-starting_Time')
    bookings = booking.objects.filter(patientid=patient).order_by('-reg_date', '-reg_time')

    # Combine and sort schedules and bookings by date and time
    history = sorted(
        list(schedules) + list(bookings),
        key=lambda x: (getattr(x, 'Date', getattr(x, 'reg_date', None)),
                       getattr(x, 'starting_Time', getattr(x, 'reg_time', None))),
        reverse=True
    )

    context = {
        'case': patient,
        'history': history,
    }
    return render(request,'main/history.html',context)
def get_patient_details(request):
    patient_id = request.GET.get('patient_id')
    try:
        patient = Patient.objects.get(id=patient_id)
        patient_data = {
            'first_name': patient.name.split()[0] if patient.name else '',
            'last_name': patient.name.split()[1] if len(patient.name.split()) > 1 else '',
            'gender': patient.gender,
            'phone': patient.phone,
            'email': patient.email,
            'dob': patient.DOB,
        }
        return JsonResponse({'success': True, 'patient_data': patient_data})
    except Patient.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Patient not found'})


def numbersession(request):
    # Get the earliest schedule for each unique combination of patient and therapist
    search_term = request.POST.get('searchitem', '')
    schedule_entries = (
        Schedule.objects
        .values('patientid', 'Therapist')  # Group by patient and therapist
        .annotate(first_schedule_id=Min('id'))  # Get the earliest schedule id per combination
    )
    per_page = int(request.GET.get('per_page', 10))
    # Extract the IDs of the first schedules from the schedule_entries
    schedule_ids = [entry['first_schedule_id'] for entry in schedule_entries]

    # Filter schedules by these unique IDs
    schedules = Schedule.objects.filter(id__in=schedule_ids)

    # Fetch all patients (for your template)
    patients = Patient.objects.all()
    if search_term:
        schedules = schedules.filter(
            Q(patientid__name__icontains=search_term) |  # Search by patient name
            Q(patientid__patient_id__icontains=search_term) |  # Search by patient ID
            Q(Therapist__name__icontains=search_term) |  # Search by therapist name
            Q(Therapy__Therapy__icontains=search_term)  # Search by therapy name
        )

    # Handle pagination
    per_page = int(request.GET.get('per_page', 10))  # Default to 10 items per page
    paginator = Paginator(schedules, per_page)
    page_number = request.GET.get('page')  # Get the page number from the URL
    page_obj = paginator.get_page(page_number)

    # Pass the schedules and patients to the template
    return render(request, 'main/numbersession.html', {'page_obj': page_obj, 'patients': patients})
def numbersessionsearch(request):
    if request.method == 'POST':
     schedule_entries = (
        Schedule.objects
        .values('patientid', 'Therapist')  # Group by patient and therapist
        .annotate(first_schedule_id=Min('id'))  # Get the earliest schedule id per combination
     )
     per_page = int(request.GET.get('per_page', 10))
    # Extract the IDs of the first schedules from the schedule_entries
     schedule_ids = [entry['first_schedule_id'] for entry in schedule_entries]

    # Filter schedules by these unique IDs
     schedules = Schedule.objects.filter(id__in=schedule_ids)
     search_term = request.POST.get('searchitem', '')
     schedules = schedules.filter(
            Q(patientid__name__icontains=search_term) |  # Search by patient name
            Q(patientid__patient_id__icontains=search_term) |  # Search by patient ID
            Q(Therapist__name__icontains=search_term) |  # Search by therapist name
            Q(Therapy__Therapy__icontains=search_term)  # Search by therapy name
     )
     page_number = request.GET.get('page', 1)  # Get the page number from the request
     paginator = Paginator(schedules, 10)  # Show 10 results per page
     page_obj = paginator.get_page(page_number)
     context = {
     'page_obj': page_obj,

     }
     return render(request, 'main/numbersession.html', context)

def update_session(request, session_id):
    session = get_object_or_404(SessionStatus, id=session_id)
    schedule = session.schedule

    if request.method == 'POST':
        if 'mark_attended' in request.POST:
            return redirect('toggle_session_status', session_id=session.id)

        session_date = request.POST.get('session_date')
        starting_time = request.POST.get('starting_Time')
        ending_time = request.POST.get('ending_Time')

        if session_date and session_date != str(session.session_date):
            session.session_date = session_date
            schedule.Date = session_date

        if starting_time and ending_time and (starting_time != str(schedule.starting_Time) or ending_time != str(schedule.ending_Time)):
            schedule.starting_Time = starting_time
            schedule.ending_Time = ending_time
            session.starting_Time = starting_time
            session.ending_Time = ending_time

        schedule.save()
        session.save()

        return redirect('view_sessions', schedule_id=schedule.id)

